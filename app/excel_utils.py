from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import os
from copy import copy

def is_colored(cell):
    fill = cell.fill
    return fill and fill.fgColor and fill.fgColor.type == 'rgb' and fill.fgColor.rgb not in ['00000000', 'FFFFFFFF']

def is_data_row(row):
    return sum(1 for cell in row if cell.value not in (None, "")) >= 2

def copy_cell_format(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.fill = copy(source_cell.fill)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format

def extract_tables_with_formatting(excel_path, output_dir, start_sheet, end_sheet):
    wb = load_workbook(excel_path)
    all_sheets = wb.sheetnames

    start_index = all_sheets.index(start_sheet)
    end_index = all_sheets.index(end_sheet)
    sheets_to_extract = all_sheets[start_index:end_index + 1]

    for sheet_name in sheets_to_extract:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows())
        table_blocks = []
        current_block = []
        block_start_indices = []
        inside_block = False

        for i, row in enumerate(rows):
            if any(is_colored(cell) for cell in row) or is_data_row(row):
                if not inside_block:
                    block_start_indices.append(i)
                    inside_block = True
                current_block.append(row)
            else:
                if inside_block:
                    table_blocks.append((block_start_indices[-1], current_block))
                    current_block = []
                    inside_block = False

        if inside_block and current_block:
            table_blocks.append((block_start_indices[-1], current_block))

        for idx, (start_row_idx, body_block) in enumerate(table_blocks):
            header_block = []
            for j in range(start_row_idx - 1, -1, -1):
                if any(cell.value for cell in rows[j]):
                    header_block.insert(0, rows[j])
                else:
                    break

            full_block = header_block + body_block
            if full_block:
                new_wb = Workbook()
                new_ws = new_wb.active
                for r_idx, row in enumerate(full_block, start=1):
                    for c_idx, cell in enumerate(row, start=1):
                        copy_cell_format(cell, new_ws.cell(row=r_idx, column=c_idx))
                new_wb.save(os.path.join(output_dir, f"{sheet_name}_{idx+1}.xlsx"))
