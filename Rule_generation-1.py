import os
from openpyxl import load_workbook

# === CONFIGURATION ===
input_folder = r'C:\Users\veerabhadra.ronad\Downloads\Input_folder\extracted_tables_with_formatting'
output_folder = r'C:\Users\veerabhadra.ronad\Downloads\Input_folder\generated_rules'  # Absolute path for consistency
os.makedirs(output_folder, exist_ok=True)

# Loop through each .xlsx file in the extracted_tables folder
for filename in os.listdir(input_folder):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(input_folder, filename)
        print(f"\n[PROCESSING] {filename}")

        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = []
            data = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                else:
                    data.append([str(cell).strip() if cell is not None else "" for cell in row])

            rule_dict = {}
            for row in data:
                for col_index, value in enumerate(row):
                    if value not in ["N", "", "None"]:
                        col_name = headers[col_index]
                        rule_dict.setdefault(col_name, set()).add(value)

            if not rule_dict:
                print(f"[SKIPPED] No rule-worthy data in {filename}")
                continue

            # Build rule text
            rule_lines = []
            rule_lines.append("AllTrue(")
            for idx, (col, values) in enumerate(rule_dict.items()):
                values_list = ', '.join([f'"{val}"' for val in sorted(values)])
                rule_lines.append(f"  AnyTrue({col}: [{values_list}])" + ("," if idx < len(rule_dict)-1 else ""))
            rule_lines.append(")")

            # Print preview in console
            print("[RULE TEXT PREVIEW]")
            print("\n".join(rule_lines))

            # Write rule text to a .txt file
            output_file_path = os.path.join(output_folder, f"{os.path.splitext(filename)[0]}_rules.txt")
            print(f"[DEBUG] Writing rule file to: {output_file_path}")

            with open(output_file_path, "w", encoding="utf-8") as f:
                f.write('\n'.join(rule_lines))

            print(f"[RULE GENERATED] {filename} → {os.path.basename(output_file_path)}")

        except Exception as e:
            print(f"[ERROR] Failed to process {filename}: {e}")

print("All rules generated.")
