import pandas as pd
import mysql.connector
import re
from openpyxl import load_workbook

# === CONFIGURATION ===
EXCEL_FILE = r"C:\Users\veerabhadra.ronad\Downloads\Input_folder\KEY-GR_PM.xlsx"
SHEET_NAME = "Product Matrix"
DB_NAME = "product_matrix_db"

# === CONNECT TO MYSQL ===
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Root"  # Change this if your MySQL password is set
)
cursor = conn.cursor()
cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME}")
cursor.execute(f"USE {DB_NAME}")
print(f"✅ Database '{DB_NAME}' ready.")

# === LOAD SHEET ===
wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

# === EXTRACT TABLE BLOCKS ===
blocks = []
current_block = []
table_name = None

for row in ws.iter_rows(values_only=True):
    if row[0] and isinstance(row[0], str) and "code" in row[0].lower():
        if current_block and table_name:
            blocks.append((table_name, pd.DataFrame(current_block).dropna(how="all")))
        current_block = []
        # Use previous row to name table (3 words from product description)
        previous_row = row
        for offset in range(1, 6):  # Look back up to 5 rows
            prev = ws[row[0].row - offset]
            desc = prev[0].value if prev and prev[0] else None
            if desc and isinstance(desc, str):
                words = re.findall(r'\w+', desc.lower())
                table_name = "_".join(words[:3]) if words else "table"
                break
        else:
            table_name = "table"
    elif any(pd.notna(cell) for cell in row):
        current_block.append(list(row))

# Add last block
if current_block and table_name:
    blocks.append((table_name, pd.DataFrame(current_block).dropna(how="all")))

# === PROCESS BLOCKS ===
for name, df in blocks:
    # Clean table name
    table_name = re.sub(r'\W+', '_', name.lower()).strip("_")
    table_name = table_name[:64]  # MySQL identifier limit

    # Use first row as header
    df.columns = df.iloc[0]
    df = df.drop(index=df.index[0]).reset_index(drop=True)
    df = df.dropna(how="all", axis=1)

    # Sanitize column names
    df.columns = [str(col).strip().lower().replace(" ", "_").replace("/", "_")[:64] for col in df.columns]

    try:
        # Drop table if exists (optional)
        cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")

        # Create table
        col_defs = ", ".join([f"`{col}` TEXT" for col in df.columns])
        cursor.execute(f"CREATE TABLE `{table_name}` ({col_defs})")
        print(f"🛠️ Created table: {table_name}")

        # Insert rows
        for _, row in df.iterrows():
            placeholders = ", ".join(["%s"] * len(row))
            col_names = ", ".join([f"`{col}`" for col in df.columns])
            cursor.execute(
                f"INSERT INTO `{table_name}` ({col_names}) VALUES ({placeholders})",
                tuple(row)
            )
        conn.commit()
        print(f"✅ Inserted {len(df)} rows into '{table_name}'.")
    except Exception as e:
        print(f"❌ Failed on table '{table_name}': {e}")

cursor.close()
conn.close()
print("✅ All done.")
