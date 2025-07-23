import os
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border

# === CONFIGURATION ===
excel_path = r"C:\Users\harshitha.b\OneDrive - SLK Software Pvt Ltd\Documents\AI Configuration\KEY-GR_PM.xlsx"
output_dir = "extracted_tables_with_formatting"
os.makedirs(output_dir, exist_ok=True)

wb = load_workbook(excel_path)
all_sheets = wb.sheetnames
start_sheet = "End_Connection"
end_sheet = "Optional_Features"

start_index = all_sheets.index(start_sheet)
end_index = all_sheets.index(end_sheet)
sheets_to_extract = all_sheets[start_index:end_index + 1]

def is_colored(cell):
    fill = cell.fill
    return (
        fill
        and fill.fgColor
        and fill.fgColor.type == 'rgb'
        and fill.fgColor.rgb not in ['00000000', 'FFFFFFFF']
    )

def is_data_row(row):
    return sum(1 for cell in row if cell.value not in (None, "")) >= 2

def copy_cell_format(source_cell, target_cell):
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.fill = copy(source_cell.fill)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format

def extract_tables_from_sheet(ws, sheet_name):
    rows = list(ws.iter_rows())
    table_blocks = []
    current_block = []
    block_start_indices = []
    inside_block = False

    for i, row in enumerate(rows):
        row_is_colored = any(is_colored(cell) for cell in row)
        row_is_data = is_data_row(row)

        if row_is_colored or row_is_data:
            if not inside_block:
                block_start_indices.append(i)
                inside_block = True
            current_block.append(row)  # Keep cell objects
        else:
            if inside_block:
                table_blocks.append((block_start_indices[-1], current_block))
                current_block = []
                inside_block = False

    if inside_block and current_block:
        table_blocks.append((block_start_indices[-1], current_block))

    # Write each table block to new Excel file with formatting
    for idx, (start_row_idx, body_block) in enumerate(table_blocks):
        header_block = []

        for j in range(start_row_idx - 1, -1, -1):
            values = [cell.value for cell in rows[j]]
            if any(values):
                header_block.insert(0, rows[j])  # Keep cell objects
            else:
                break

        full_block = header_block + body_block
        if full_block:
            new_wb = Workbook()
            new_ws = new_wb.active

            for r_idx, row in enumerate(full_block, start=1):
                for c_idx, cell in enumerate(row, start=1):
                    copy_cell_format(cell, new_ws.cell(row=r_idx, column=c_idx))

            filename = f"{sheet_name}_{idx + 1}.xlsx"
            save_path = os.path.join(output_dir, filename)
            new_wb.save(save_path)
            print(f"✅ Saved: {filename}")

# 🔁 Run for selected sheets
for sheet_name in sheets_to_extract:
    print(f"Processing sheet: {sheet_name}")
    extract_tables_from_sheet(wb[sheet_name], sheet_name)

print("🎉 Done. All tables saved with formatting.") 